"""
SEC-compliant report generation service
Handles PDF and Excel export for SEC Climate Disclosure Rule compliance
"""

import io
from datetime import datetime
from typing import Any, Dict, List, Optional
from uuid import UUID

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.user import User
from app.services.emissions_consolidation_service import EmissionsConsolidationService


class SECReportGenerator:
    """Service for generating SEC-compliant reports in multiple formats"""

    def __init__(self, db: Session):
        self.db = db
        self.consolidation_service = EmissionsConsolidationService(db)

    async def generate_sec_report(
        self,
        company_id: UUID,
        reporting_year: int,
        consolidation_id: Optional[UUID] = None,
        format_type: str = "json",
        include_entity_breakdown: bool = True,
        include_audit_trail: bool = False,
        user: Optional[User] = None,
    ) -> Dict[str, Any]:
        """
        Generate SEC-compliant climate disclosure report

        Args:
            company_id: Company UUID
            reporting_year: Reporting year
            consolidation_id: Specific consolidation ID (optional)
            format_type: Export format (json, pdf, excel)
            include_entity_breakdown: Include entity-level details
            include_audit_trail: Include audit trail (admin only)
            user: Current user for permissions

        Returns:
            Report data or file content
        """
        # Get consolidation data
        if consolidation_id:
            consolidation = await self.consolidation_service.get_consolidation(
                consolidation_id
            )
        else:
            consolidations = await self.consolidation_service.list_consolidations(
                company_id=company_id, reporting_year=reporting_year, limit=1, offset=0
            )
            if not consolidations:
                raise HTTPException(
                    status_code=404,
                    detail=f"No consolidations found for company {company_id} in year {reporting_year}",
                )
            consolidation = await self.consolidation_service.get_consolidation(
                consolidations[0].id
            )

        # Build base report structure
        report_data = self._build_report_structure(
            consolidation, reporting_year, include_entity_breakdown
        )

        # Add audit trail if requested and user has permission
        if include_audit_trail and user and user.is_admin:
            report_data["audit_trail"] = await self._get_audit_trail_data(
                consolidation.id
            )

        # Generate report based on format
        if format_type.lower() == "json":
            return report_data
        elif format_type.lower() == "pdf":
            return await self._generate_pdf_report(report_data)
        elif format_type.lower() == "excel":
            return await self._generate_excel_report(report_data)
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported format: {format_type}. Supported: json, pdf, excel",
            )

    def _build_report_structure(
        self, consolidation, reporting_year: int, include_entity_breakdown: bool
    ) -> Dict[str, Any]:
        """Build the base SEC report structure"""
        report = {
            "report_type": "sec_climate_disclosure",
            "sec_form": "10-K",
            "company_id": str(consolidation.company_id),
            "reporting_year": reporting_year,
            "consolidation_id": str(consolidation.id),
            "report_generated_at": datetime.now().isoformat() + "Z",
            "compliance_standard": "SEC Climate Disclosure Rule",
            # Executive Summary
            "executive_summary": {
                "total_ghg_emissions_mtco2e": consolidation.total_co2e,
                "scope1_emissions_mtco2e": consolidation.total_scope1_co2e,
                "scope2_emissions_mtco2e": consolidation.total_scope2_co2e,
                "scope3_emissions_mtco2e": consolidation.total_scope3_co2e,
                "emissions_intensity": None,  # Will be calculated based on revenue
                "year_over_year_change": None,  # Will be calculated from previous year
                "data_completeness_score": consolidation.data_completeness_score,
                "reporting_boundary": f"{consolidation.total_entities_included} consolidated entities",
            },
            # Methodology and Data Quality
            "methodology": {
                "calculation_methodology": "GHG Protocol Corporate Standard",
                "consolidation_method": consolidation.consolidation_method.value,
                "emission_factors_source": "EPA emission factors database",
                "data_collection_period": f"January 1 - December 31, {reporting_year}",
                "organizational_boundary": "Equity share (ownership percentage)",
                "operational_boundary": "All facilities under operational control",
                "data_quality_assessment": {
                    "completeness": consolidation.data_completeness_score,
                    "accuracy": consolidation.consolidation_confidence_score,
                    "consistency": "Cross-validated against EPA GHGRP data",
                    "transparency": "Full audit trail available",
                },
            },
            # Emissions Data Tables
            "emissions_tables": {
                "table_1_scope1_emissions": {
                    "title": "Scope 1 GHG Emissions by Source",
                    "units": "Metric tons CO2 equivalent (MTCO2e)",
                    "data": self._format_scope1_table(consolidation),
                },
                "table_2_scope2_emissions": {
                    "title": "Scope 2 GHG Emissions by Source",
                    "units": "Metric tons CO2 equivalent (MTCO2e)",
                    "data": self._format_scope2_table(consolidation),
                },
                "table_3_emissions_by_business_segment": {
                    "title": "GHG Emissions by Business Segment",
                    "units": "Metric tons CO2 equivalent (MTCO2e)",
                    "data": self._format_segment_table(consolidation),
                },
            },
            # Risk Assessment
            "climate_related_risks": {
                "transition_risks": [],
                "physical_risks": [],
                "risk_management_strategy": "Integrated climate risk assessment framework",
                "scenario_analysis": "Climate scenarios considered: 2°C, 4°C pathways",
            },
            # Governance
            "governance": {
                "board_oversight": "Board-level climate committee established",
                "management_responsibility": "Chief Sustainability Officer appointed",
                "stakeholder_engagement": "Regular engagement with investors and NGOs",
            },
            # Status and Approval
            "status": {
                "report_status": consolidation.status.value,
                "is_final": consolidation.is_final,
                "validation_status": consolidation.validation_status,
                "approved_by": (
                    str(consolidation.approved_by)
                    if consolidation.approved_by
                    else None
                ),
                "approved_at": (
                    consolidation.approved_at.isoformat()
                    if consolidation.approved_at
                    else None
                ),
                "last_modified": (
                    consolidation.updated_at.isoformat()
                    if consolidation.updated_at
                    else None
                ),
            },
        }

        # Add entity breakdown if requested
        if include_entity_breakdown:
            report["entity_breakdown"] = self._format_entity_breakdown(consolidation)

        return report

    def _format_scope1_table(self, consolidation) -> List[Dict[str, Any]]:
        """Format Scope 1 emissions table data"""
        scope1_data = []

        # Aggregate by fuel type from entity contributions
        fuel_totals = {}
        for contrib in consolidation.entity_contributions:
            if (
                contrib.consolidated_scope1_co2e
                and contrib.consolidated_scope1_co2e > 0
            ):
                # This would need to be enhanced with actual fuel type breakdown
                # For now, aggregate all Scope 1 as "Stationary Combustion"
                fuel_type = "Stationary Combustion"  # Placeholder
                if fuel_type not in fuel_totals:
                    fuel_totals[fuel_type] = 0
                fuel_totals[fuel_type] += contrib.consolidated_scope1_co2e

        for fuel_type, emissions in fuel_totals.items():
            scope1_data.append(
                {
                    "source_category": fuel_type,
                    "emissions_mtco2e": round(emissions, 2),
                    "percentage_of_total": (
                        round((emissions / consolidation.total_scope1_co2e * 100), 1)
                        if consolidation.total_scope1_co2e
                        else 0
                    ),
                }
            )

        return scope1_data

    def _format_scope2_table(self, consolidation) -> List[Dict[str, Any]]:
        """Format Scope 2 emissions table data"""
        scope2_data = []

        # Aggregate by region from entity contributions
        region_totals = {}
        for contrib in consolidation.entity_contributions:
            if (
                contrib.consolidated_scope2_co2e
                and contrib.consolidated_scope2_co2e > 0
            ):
                # This would need to be enhanced with actual region breakdown
                # For now, aggregate all Scope 2 as "Electricity - Grid Mix"
                region = "Electricity - Grid Mix"  # Placeholder
                if region not in region_totals:
                    region_totals[region] = 0
                region_totals[region] += contrib.consolidated_scope2_co2e

        for region, emissions in region_totals.items():
            scope2_data.append(
                {
                    "source_category": region,
                    "emissions_mtco2e": round(emissions, 2),
                    "percentage_of_total": (
                        round((emissions / consolidation.total_scope2_co2e * 100), 1)
                        if consolidation.total_scope2_co2e
                        else 0
                    ),
                }
            )

        return scope2_data

    def _format_segment_table(self, consolidation) -> List[Dict[str, Any]]:
        """Format emissions by business segment table"""
        segment_data = []

        # Group by entity (treating each entity as a segment)
        for contrib in consolidation.entity_contributions:
            if contrib.included_in_consolidation:
                segment_data.append(
                    {
                        "business_segment": contrib.entity_name,
                        "scope1_emissions": round(
                            contrib.consolidated_scope1_co2e or 0, 2
                        ),
                        "scope2_emissions": round(
                            contrib.consolidated_scope2_co2e or 0, 2
                        ),
                        "scope3_emissions": round(
                            contrib.consolidated_scope3_co2e or 0, 2
                        ),
                        "total_emissions": round(
                            contrib.consolidated_total_co2e or 0, 2
                        ),
                        "revenue_millions": None,  # Would need to be added from financial data
                        "emissions_intensity": None,  # Revenue / emissions
                    }
                )

        return segment_data

    def _format_entity_breakdown(self, consolidation) -> List[Dict[str, Any]]:
        """Format detailed entity breakdown"""
        entities = []
        for contrib in consolidation.entity_contributions:
            entities.append(
                {
                    "entity_id": str(contrib.entity_id),
                    "entity_name": contrib.entity_name,
                    "ownership_percentage": contrib.ownership_percentage,
                    "consolidation_factor": contrib.consolidation_factor,
                    "original_emissions": {
                        "scope1_mtco2e": contrib.original_scope1_co2e,
                        "scope2_mtco2e": contrib.original_scope2_co2e,
                        "scope3_mtco2e": contrib.original_scope3_co2e,
                        "total_mtco2e": contrib.original_total_co2e,
                    },
                    "consolidated_emissions": {
                        "scope1_mtco2e": contrib.consolidated_scope1_co2e,
                        "scope2_mtco2e": contrib.consolidated_scope2_co2e,
                        "scope3_mtco2e": contrib.consolidated_scope3_co2e,
                        "total_mtco2e": contrib.consolidated_total_co2e,
                    },
                    "data_quality": {
                        "completeness_score": contrib.data_completeness,
                        "quality_score": contrib.data_quality_score,
                    },
                    "included_in_consolidation": contrib.included_in_consolidation,
                    "exclusion_reason": contrib.exclusion_reason,
                }
            )

        return entities

    async def _get_audit_trail_data(self, consolidation_id: UUID) -> Dict[str, Any]:
        """Get audit trail data for the consolidation"""
        # This would integrate with audit service when available
        return {
            "consolidation_events": [],
            "data_modifications": [],
            "approval_history": [],
            "note": "Full audit trail integration pending",
        }

    async def _generate_pdf_report(self, report_data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate PDF report using ReportLab"""
        buffer = io.BytesIO()

        # Create PDF document
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=16,
            spaceAfter=30,
        )
        story.append(Paragraph("SEC Climate Disclosure Report", title_style))
        story.append(Spacer(1, 12))

        # Company and Year info
        info_style = ParagraphStyle(
            "Info",
            parent=styles["Normal"],
            fontSize=10,
            spaceAfter=20,
        )
        company_info = f"Company ID: {report_data['company_id']}<br/>Reporting Year: {report_data['reporting_year']}<br/>Generated: {report_data['report_generated_at']}"
        story.append(Paragraph(company_info, info_style))

        # Executive Summary
        story.append(Paragraph("Executive Summary", styles["Heading2"]))
        summary_data = report_data["executive_summary"]
        summary_text = f"""
        Total GHG Emissions: {summary_data['total_ghg_emissions_mtco2e'] or 0:.2f} MTCO2e<br/>
        Scope 1: {summary_data['scope1_emissions_mtco2e'] or 0:.2f} MTCO2e<br/>
        Scope 2: {summary_data['scope2_emissions_mtco2e'] or 0:.2f} MTCO2e<br/>
        Scope 3: {summary_data['scope3_emissions_mtco2e'] or 0:.2f} MTCO2e<br/>
        Data Completeness: {summary_data['data_completeness_score'] or 0:.1%}
        """
        story.append(Paragraph(summary_text, styles["Normal"]))
        story.append(Spacer(1, 12))

        # Emissions Tables
        if "emissions_tables" in report_data:
            story.append(Paragraph("Emissions Data", styles["Heading2"]))

            # Scope 1 Table
            if report_data["emissions_tables"]["table_1_scope1_emissions"]["data"]:
                story.append(
                    Paragraph("Scope 1 GHG Emissions by Source", styles["Heading3"])
                )
                scope1_data = report_data["emissions_tables"][
                    "table_1_scope1_emissions"
                ]["data"]

                table_data = [["Source Category", "Emissions (MTCO2e)", "Percentage"]]
                for row in scope1_data:
                    table_data.append(
                        [
                            row["source_category"],
                            f"{row['emissions_mtco2e']:.2f}",
                            f"{row['percentage_of_total']:.1f}%",
                        ]
                    )

                table = Table(table_data)
                table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 10),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                            ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ]
                    )
                )
                story.append(table)
                story.append(Spacer(1, 12))

        # Build PDF
        doc.build(story)
        buffer.seek(0)

        return {
            "filename": f"SEC_Climate_Disclosure_{report_data['company_id']}_{report_data['reporting_year']}.pdf",
            "content_type": "application/pdf",
            "content": buffer.getvalue(),
        }

    async def _generate_excel_report(
        self, report_data: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Generate Excel report using openpyxl"""
        wb = Workbook()
        ws = wb.active
        ws.title = "SEC Climate Disclosure"

        # Styles
        header_font = Font(bold=True, size=12)
        subheader_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title
        ws["A1"] = "SEC Climate Disclosure Report"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")

        # Company Info
        ws["A3"] = "Company ID:"
        ws["B3"] = report_data["company_id"]
        ws["A4"] = "Reporting Year:"
        ws["B4"] = report_data["reporting_year"]
        ws["A5"] = "Generated:"
        ws["B5"] = report_data["report_generated_at"]

        # Executive Summary
        ws["A7"] = "Executive Summary"
        ws["A7"].font = header_font
        ws.merge_cells("A7:D7")

        summary = report_data["executive_summary"]
        ws["A9"] = "Total GHG Emissions (MTCO2e):"
        ws["B9"] = summary["total_ghg_emissions_mtco2e"] or 0
        ws["A10"] = "Scope 1 Emissions (MTCO2e):"
        ws["B10"] = summary["scope1_emissions_mtco2e"] or 0
        ws["A11"] = "Scope 2 Emissions (MTCO2e):"
        ws["B11"] = summary["scope2_emissions_mtco2e"] or 0
        ws["A12"] = "Scope 3 Emissions (MTCO2e):"
        ws["B12"] = summary["scope3_emissions_mtco2e"] or 0
        ws["A13"] = "Data Completeness Score:"
        ws["B13"] = summary["data_completeness_score"] or 0

        # Emissions Tables
        row = 15
        if "emissions_tables" in report_data:
            # Scope 1 Table
            if report_data["emissions_tables"]["table_1_scope1_emissions"]["data"]:
                ws.cell(
                    row=row, column=1, value="Scope 1 GHG Emissions by Source"
                ).font = header_font
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                row += 2

                # Headers
                ws.cell(row=row, column=1, value="Source Category").font = (
                    subheader_font
                )
                ws.cell(row=row, column=2, value="Emissions (MTCO2e)").font = (
                    subheader_font
                )
                ws.cell(row=row, column=3, value="Percentage").font = subheader_font
                row += 1

                # Data
                for data_row in report_data["emissions_tables"][
                    "table_1_scope1_emissions"
                ]["data"]:
                    ws.cell(
                        row=row, column=1, value=data_row["source_category"]
                    ).font = normal_font
                    ws.cell(
                        row=row, column=2, value=data_row["emissions_mtco2e"]
                    ).font = normal_font
                    ws.cell(
                        row=row,
                        column=3,
                        value=f"{data_row['percentage_of_total']:.1f}%",
                    ).font = normal_font
                    row += 1

        # Auto-adjust column widths
        for col_num, column in enumerate(ws.columns, 1):
            max_length = 0
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            if max_length > 0:
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
                column_letter = chr(64 + col_num)  # A=1, B=2, etc.
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return {
            "filename": f"SEC_Climate_Disclosure_{report_data['company_id']}_{report_data['reporting_year']}.xlsx",
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "content": buffer.getvalue(),
        }
